from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.entity_catalog import normalize_entity_text


WORKBOOK_PATH = Path(__file__).resolve().parents[1] / "BD_Entidades_Destino_Defiendo_Colombia.xlsx"


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _split_contacts(value: Any) -> list[str]:
    text = _clean(value)
    if not text:
        return []
    parts = re.split(r"\s*/\s*|\n+", text)
    return [part.strip() for part in parts if part and part.strip()]


def _pick_emails(values: list[str]) -> list[str]:
    return [value for value in values if "@" in value]


def _pick_phones(values: list[str]) -> list[str]:
    return [value for value in values if "@" not in value and not value.lower().startswith("http")]


def _extract_aliases(name: str) -> list[str]:
    aliases: list[str] = []
    compact = re.sub(r"\s+", " ", name).strip()
    aliases.append(compact)

    parentheses = re.findall(r"\(([^)]+)\)", compact)
    aliases.extend(parentheses)

    without_parentheses = re.sub(r"\s*\([^)]*\)", "", compact).strip()
    if without_parentheses and without_parentheses != compact:
        aliases.append(without_parentheses)

    if " - " in compact:
        aliases.extend(part.strip() for part in compact.split(" - ") if part.strip())

    deduped: list[str] = []
    seen: set[str] = set()
    for alias in aliases:
        key = normalize_entity_text(alias)
        if not key or key in seen:
            continue
        seen.add(key)
        deduped.append(alias)
    return deduped


def _index_entity(entity: dict[str, Any]) -> dict[str, Any]:
    item = dict(entity)
    item["name_search"] = normalize_entity_text(item["name"])
    item["aliases_search"] = [normalize_entity_text(alias) for alias in item.get("aliases", [])]
    return item


def _routing(channel: str, contact: str | None, notes: str | None, automatable: bool = True) -> dict[str, Any]:
    return {
        "channel": channel,
        "contact": contact,
        "automatable": automatable,
        "genera_radicado": True,
        "response_window": None,
        "notes": notes,
    }


def _load_eps_sheet(ws: Any) -> list[dict[str, Any]]:
    entities: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = _clean(row[0])
        if not name:
            continue
        phones = _split_contacts(row[4])
        notes = _clean(row[6])
        website = _clean(row[2])
        emails = _pick_emails(_split_contacts(row[3]))
        entities.append(
            _index_entity(
                {
                    "name": name,
                    "aliases": _extract_aliases(name),
                    "type": "eps",
                    "sector": "salud",
                    "phone": phones[0] if phones else None,
                    "phones": phones,
                    "pqrs_emails": emails,
                    "website": website,
                    "superintendence": "Supersalud",
                    "source": "excel_maestro",
                    "verified": True,
                    "routing_channels": [
                        _routing("portal_pqrs", website, notes),
                        *[_routing("email", email, notes) for email in emails],
                        *[_routing("telefono", phone, notes, automatable=False) for phone in phones],
                    ],
                }
            )
        )
    return entities


def _load_banks_sheet(ws: Any) -> list[dict[str, Any]]:
    entities: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = _clean(row[0])
        if not name:
            continue
        phones = _split_contacts(row[4])
        notes = _clean(row[6])
        website = _clean(row[2])
        emails = _pick_emails(_split_contacts(row[3]))
        defender_emails = _pick_emails(_split_contacts(row[5]))
        entities.append(
            _index_entity(
                {
                    "name": name,
                    "aliases": _extract_aliases(name),
                    "type": "banco",
                    "sector": "financiero",
                    "phone": phones[0] if phones else None,
                    "phones": phones,
                    "pqrs_emails": emails,
                    "notification_emails": defender_emails,
                    "website": website,
                    "superintendence": "Superfinanciera",
                    "source": "excel_maestro",
                    "verified": True,
                    "routing_channels": [
                        _routing("portal_pqrs", website, notes),
                        *[_routing("email", email, notes) for email in emails],
                        *[_routing("defensor_consumidor", email, notes, automatable=False) for email in defender_emails],
                        *[_routing("telefono", phone, notes, automatable=False) for phone in phones],
                    ],
                }
            )
        )
    return entities


def _service_type(service_name: str | None) -> tuple[str, str, str]:
    text = normalize_entity_text(service_name)
    if any(token in text for token in ("internet", "telefon", "movil", "celular", "telecom")):
        return ("telecom", "telecomunicaciones", "CRC")
    return ("servicio_publico", service_name or "servicios_publicos", "Superservicios")


def _load_services_sheet(ws: Any) -> list[dict[str, Any]]:
    entities: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = _clean(row[0])
        if not name:
            continue
        type_name, sector, oversight = _service_type(_clean(row[1]))
        phones = _split_contacts(row[5])
        notes = _clean(row[7])
        website = _clean(row[3])
        emails = _pick_emails(_split_contacts(row[4]))
        entities.append(
            _index_entity(
                {
                    "name": name,
                    "aliases": _extract_aliases(name),
                    "type": type_name,
                    "sector": sector,
                    "city": _clean(row[2]),
                    "phone": phones[0] if phones else None,
                    "phones": phones,
                    "pqrs_emails": emails,
                    "website": website,
                    "superintendence": oversight,
                    "source": "excel_maestro",
                    "verified": True,
                    "routing_channels": [
                        _routing("portal_pqrs", website, notes),
                        *[_routing("email", email, notes) for email in emails],
                        *[_routing("telefono", phone, notes, automatable=False) for phone in phones],
                    ],
                }
            )
        )
    return entities


def _load_public_sheet(ws: Any) -> list[dict[str, Any]]:
    entities: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = _clean(row[0])
        if not name:
            continue
        phones = _split_contacts(row[4])
        official_emails = _pick_emails(_split_contacts(row[2]))
        alternate_emails = _pick_emails(_split_contacts(row[3]))
        website = _clean(row[5])
        entities.append(
            _index_entity(
                {
                    "name": name,
                    "aliases": _extract_aliases(name),
                    "type": "entidad_publica",
                    "sector": _clean(row[1]) or "gobierno",
                    "phone": phones[0] if phones else None,
                    "phones": phones,
                    "pqrs_emails": official_emails,
                    "notification_emails": alternate_emails,
                    "website": website,
                    "source": "excel_maestro",
                    "verified": True,
                    "routing_channels": [
                        _routing("portal_pqrs", website, "Entidad publica con canal oficial"),
                        *[_routing("email", email, "Correo oficial") for email in official_emails],
                        *[_routing("email_alterno", email, "Correo alternativo", automatable=False) for email in alternate_emails],
                        *[_routing("telefono", phone, "PBX o linea oficial", automatable=False) for phone in phones],
                    ],
                }
            )
        )
    return entities


@lru_cache(maxsize=1)
def load_excel_entities() -> list[dict[str, Any]]:
    if not WORKBOOK_PATH.exists():
        return []

    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        entities: list[dict[str, Any]] = []
        entities.extend(_load_eps_sheet(workbook["📧 EPS Colombia"]))
        entities.extend(_load_banks_sheet(workbook["🏦 Bancos Colombia"]))
        entities.extend(_load_services_sheet(workbook["⚡ Empresas SP"]))
        entities.extend(_load_public_sheet(workbook["🏛️ Entidades Públicas"]))
    finally:
        workbook.close()
    return entities


def search_excel_entities(query: str, limit: int = 8) -> list[dict[str, Any]]:
    normalized_query = normalize_entity_text(query)
    if len(normalized_query) < 2:
        return []

    matches: list[tuple[int, dict[str, Any]]] = []
    for entity in load_excel_entities():
        candidates = [entity["name_search"], *entity.get("aliases_search", [])]
        if normalized_query in candidates:
            score = 0
        elif any(candidate.startswith(normalized_query) for candidate in candidates):
            score = 1
        elif any(normalized_query in candidate for candidate in candidates):
            score = 2
        else:
            continue
        matches.append((score, entity))

    matches.sort(key=lambda item: (item[0], item[1]["name"]))
    return [dict(entity) for _, entity in matches[:limit]]
